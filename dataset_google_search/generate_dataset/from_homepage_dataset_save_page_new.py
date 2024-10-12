# -*- coding: utf-8 -*-

# @Time : 2023/01/09 20:42
# @Author : luff543, Justine yeh
# @Email : luff543@gmail.com, justine8811@gmail.com
# @File : simple_markup_tool.py
# @Software: PyCharm
import math
import argparse
from engines.configure import Configure
from engines.utils.io import fold_check, make_directory
from engines.utils.logger import get_logger
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import codecs
import time
from PIL import Image
# from bs4 import BeautifulSoup

import bs4

Image.LOAD_TRUNCATED_IMAGES = True
# Image.MAX_IMAGE_PIXELS = None
# Image.MAX_IMAGE_PIXELS = 1920 * 10 * 1024 / 4 / 3
from io import BytesIO
import cv2

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.common.exceptions import TimeoutException
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
import glob
import os.path
from os import path
from natsort import natsorted
import pickle
import subprocess
import shutil
import traceback
from time import sleep
# urls = ["http://encounter.org.tw/news/", "http://springpoolglass.com/news/", "https://www.n2.org.tw/enter/news",
#             "https://www.toeic.com.tw/info/article/newsletter/", "https://www.toeic.com.tw/info/notice/announcement/"]
#
# urls = ["http://encounter.org.tw/news/", "http://springpoolglass.com/news/"]

IMGNAME_2_URL_PICKLE_NAME = "./dataset_google_search/generate_dataset/imgname_2_url.pkl"


def change_name(original_name):
    name = original_name.replace("http://", "")
    name = name.replace("https://", "")
    if name.endswith("/"):
        name = name[0:-1]
    name = name.replace("/", "__")

    name = name.replace("?", "_")
    name = name.replace("&", "_")

    return name

def read_urls(path, sheet_name="google_search_rank_train"):
    names = []
    urls = []

    df = pd.read_excel(path, sheet_name=sheet_name, header=0,
                       converters={"id": int, "domain_page": str})

    for i in range(2):
        names.append("")
        urls.append("")

    for index, row in df.iterrows():
        id = row["id"]
        print(f'id: {id}')
        if math.isnan(id):
            continue
        homepage = row["domain_page"]
        # cache = row["Cache"]
        name = change_name(homepage)
        # name = homepage.replace("http://", "")
        # name = name.replace("https://", "")
        # if name.endswith("/"):
        #     name = name[0:-1]
        # name = name.replace("/", "__")

        print(f'homepage: {homepage}, name: {name}')

        names[id] = name
        # urls[id] = cache
        urls[id] = homepage

    return names, urls


def init_driver():
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--incognito')
    options.add_argument('--start-maximized')
    options.add_argument('--disable-gpu')
    options.add_argument("disable-features=NetworkService")

    service = ChromeService(ChromeDriverManager().install())
    # options.binary_location = "C:/Users/Luff/.wdm/drivers/chromedriver/win32/108.0.5359/chromedriver.exe"
    options.binary_location = "C:/Program Files/Google/Chrome/Application/chrome.exe"
    driver = webdriver.Chrome(service=service, options=options)
    driver.maximize_window()
    # driver.set_page_load_timeout(120)
    # driver.set_page_load_timeout(180)
    driver.set_page_load_timeout(240)

    return driver


def scroll(driver):
    driver.execute_script(""" 
        (function () { 
            var y = document.body.scrollTop; 
            var step = 100; 
            window.scroll(0, y); 
            function f() { 
                if (y < document.body.scrollHeight) { 
                    y += step; 
                    window.scroll(0, y); 
                    setTimeout(f, 50); 
                }
                else { 
                    window.scroll(0, y); 
                    document.title += "scroll-done"; 
                } 
            } 
            setTimeout(f, 1000); 
        })(); 
        """)


def open_url(driver, url, selenium_html_name_path, singlefile_html_name_path, singlefile_bs4_html_name_path,
             singlefile_html_with_hidden_elements_name_path, singlefile_with_hidden_elements_bs4_html_name_path,
             html_path, with_hidden_elements_html_path, image_path):
    try:
        if not os.path.exists(selenium_html_name_path):
            driver.set_window_size(1920, 1080)
            driver.get(url)
            # save_page(driver, html_path)
            try:
                a1 = driver.switch_to.alert  # 通過switch_to.alert切換到alert
                sleep(1)
                print(f"alert:{a1.text}")  # text屬性輸出alert的文本
                a1.accept()  # alert“确认”
                sleep(1)
            except Exception:
                print('no alert')
            finally:
                save_page_by_selenium(driver, selenium_html_name_path)
                save_screenshot(driver, image_path)
    except TimeoutException as e:
        logger.error(f"url: {url}, selenium TimeoutException has been thrown. message: {e}")
        # driver.close()
        driver = init_driver()
    except Exception as e:
        logger.error(f"url: {url}, selenium Exception has been thrown. message: {e}")
        # driver.close()
        driver = init_driver()

    try:
        if not os.path.exists(singlefile_html_name_path):
            save_page_by_singlefile(url, singlefile_html_name_path)
        # shutil.copyfile(singlefile_html_name_path, html_path)
    except Exception as e:
        logger.error(f"url: {url}, singlefile Exception has been thrown. message: {e}")

    try:
        if not os.path.exists(singlefile_bs4_html_name_path) and os.path.exists(singlefile_html_name_path):
            bs4_html = read_html_by_bs4(singlefile_html_name_path)
            # save_page_by_singlefile(url, singlefile_bs4_html_name_path)
            with codecs.open(singlefile_bs4_html_name_path, "w", "utf-8") as file:
                file.write(bs4_html)
    except Exception as e:
        logger.error(f"url: {url}, singlefile_bs4 Exception has been thrown. message: {e}")
        traceback.print_exc()
    try:
        if not os.path.exists(html_path) and os.path.exists(singlefile_html_name_path):
            remove_cach_element_normalizeurl(singlefile_html_name_path, html_path, url)
    except Exception as e:
        logger.error(f"url: {url}, Exception has been thrown. message: {e}")
        traceback.print_exc()

    ## new add
    try:
        if not os.path.exists(singlefile_html_with_hidden_elements_name_path):
            save_page_by_singlefile(url, singlefile_html_with_hidden_elements_name_path, True)
        # shutil.copyfile(singlefile_html_name_path, html_path)
    except Exception as e:
        logger.error(f"url: {url}, singlefile_with_hidden_elements Exception has been thrown. message: {e}")

    try:
        if not os.path.exists(singlefile_with_hidden_elements_bs4_html_name_path) and \
                os.path.exists(singlefile_html_with_hidden_elements_name_path):
            hidden_element_bs4_html = read_html_by_bs4(singlefile_html_with_hidden_elements_name_path)
            # save_page_by_singlefile(url, singlefile_bs4_html_name_path)
            with codecs.open(singlefile_with_hidden_elements_bs4_html_name_path, "w", "utf-8") as file:
                file.write(hidden_element_bs4_html)
    except Exception as e:
        logger.error(f"url: {url}, singlefile_with_hidden_elements_bs4 Exception has been thrown. message: {e}")
        traceback.print_exc()

    try:
        if not os.path.exists(with_hidden_elements_html_path) and os.path.exists(singlefile_html_with_hidden_elements_name_path):
            remove_cach_element_normalizeurl(singlefile_html_with_hidden_elements_name_path, with_hidden_elements_html_path, url)
    except Exception as e:
        logger.error(f"url: {url}, hidden element Exception has been thrown. message: {e}")
        traceback.print_exc()
    return driver


def find_nth(haystack, needle, n):
    start = haystack.find(needle)
    while start >= 0 and n > 1:
        start = haystack.find(needle, start + len(needle))
        n -= 1
    return start


def mask_prefix_string(cache_url):
    cache_prefix_index = find_nth(cache_url, "http", 2)
    cache_prefix = cache_url[0:cache_prefix_index]
    print('cache_prefix',cache_prefix)
    cache_relevance_prefix_index = find_nth(cache_prefix, "/web/", 1)
    cache_relevance_prefix = cache_prefix[cache_relevance_prefix_index:]
    print('cache_relevance_prefix',cache_relevance_prefix)

    return cache_prefix, cache_relevance_prefix

def read_html_by_bs4(html_path):
    f = codecs.open(html_path, "r", "utf-8")
    html = f.read()
    f.close()

    soup = bs4.BeautifulSoup(html, 'html5lib')
    soup_html = str(soup)

    return soup_html


def custom_replace(content, target, replace_symbol):
    result = ""
    lines = []
    for line in content.splitlines():
        replace_line = line.replace(target, replace_symbol)
        replace_line = replace_line.replace("\n", "")
        lines.append(replace_line)

    result = "\n".join(lines)

    return result


def find_childrens(soup, tag_name, filter_condiion):
    target_element = soup.find(tag_name, filter_condiion)
    childrens = getattr(target_element, "children", iter([]))

    return childrens

def find_tags(soup, tag_name, filter_condiion):
    target_element = soup.find(tag_name, filter_condiion)
    # childrens = getattr(target_element, "children", iter([]))

    return target_element
def remove_cach_element_normalizeurl(html_path, output_html_Path, ref_url):
    f = codecs.open(html_path, "r", "utf-8")
    original_html = f.read()
    f.close()

    soup = bs4.BeautifulSoup(original_html, 'html5lib')

    # for current_tag in soup.find("iframe", {"id": "bk-global-iframe"}).children:
    #     if isinstance(current_tag, bs4.element.Tag):
    #         current_tag.decompose()

    tag_name = "div"
    filter_condiion = {"id": "wm-ipp-base"}
    target_tag = find_tags(soup, tag_name, filter_condiion)
    if isinstance(target_tag, bs4.element.Tag):
        target_tag.decompose()
    # for current_tag in find_childrens(soup, tag_name, filter_condiion):
    # # for current_tag in soup.find("div", {"id": "wm-ipp-base"}).children:
    #     if isinstance(current_tag, bs4.element.Tag):
    #         current_tag.decompose()
    # for current_tag in soup.find("div", {"id": "wm-ipp-base"}).children:
    #     if isinstance(current_tag, bs4.element.Tag):
    #         current_tag.decompose()

    tag_name = "div"
    filter_condiion = {"id": "wm-ipp-print"}
    target_tag = find_tags(soup, tag_name, filter_condiion)
    if isinstance(target_tag, bs4.element.Tag):
        target_tag.decompose()
    # for current_tag in find_childrens(soup, tag_name, filter_condiion):
    # # for current_tag in soup.find("div", {"id": "wm-ipp-print"}).children:
    #     if isinstance(current_tag, bs4.element.Tag):
    #         current_tag.decompose()

    tag_name = "div"
    filter_condiion = {"class": "sf-hidden"}
    target_tag = find_tags(soup, tag_name, filter_condiion)
    if isinstance(target_tag, bs4.element.Tag):
        target_tag.decompose()
    # for current_tag in find_childrens(soup, tag_name, filter_condiion):
    # # for current_tag in soup.find("div", {"class": "sf-hidden"}).children:
    #     if isinstance(current_tag, bs4.element.Tag):
    #         current_tag.decompose()  # 清除標籤 (soup.tag.decompose)

    soup_html = str(soup)
    # cache_prefix, cache_relevance_prefix = mask_prefix_string(ref_url)
    # soup_html = soup_html.replace(cache_url, "")
    # soup_html = soup_html.replace(cache_relevance_prefix, "")
    # soup_html = custom_replace(soup_html, cache_prefix, "")
    # soup_html = custom_replace(soup_html, cache_relevance_prefix, "")
    # print(f"cache_prefix: {cache_prefix}, cache_relevance_prefix:{cache_relevance_prefix}")

    with open(output_html_Path, "w", encoding='utf-8') as file:
        file.write(soup_html)


def save_screenshot(driver, file_name, custom_width=1920):
    scroll(driver)
    height, width = scroll_down(driver)
    if custom_width != None:
        width = custom_width
    driver.set_window_size(width, height)
    img_binary = driver.get_screenshot_as_png()
    img = Image.open(BytesIO(img_binary))
    img.save(file_name)
    logger.info(" screenshot saved ")


def convert_png_to_jpg(original_path, dest_path):
    compress_img_by_cv(original_path, dest_path)
    # try:
    #     compress_img_by_cv(original_path, dest_path)
    # except Exception as e:
    #     logger.info(f"Exception has been thrown message: {e}")

    # original_img = Image.open(original_path)
    # rgb_img = original_img.convert('RGB')
    # rgb_img.save(dest_path)


def compress_img_by_cv(img_path, output_path, compress_rate=0.8):
    img = cv2.imread(img_path)
    # img_resize = cv2.resize(img, (int(heigh * compress_rate), int(width * compress_rate)),
    #                         interpolation=cv2.INTER_AREA)
    resize_img = image_resize(img, width=1920, height=None)
    cv2.imwrite(output_path, resize_img, [int(cv2.IMWRITE_JPEG_QUALITY), 95])
    logger.info(f"壓縮: {img_path} 輸出路徑: {output_path}")


def image_resize(image, width=None, height=None, inter=cv2.INTER_AREA):
    dim = None
    (h, w) = image.shape[:2]

    if width is None and height is None:
        return image

    if width is None:
        r = height / float(h)
        dim = (int(w * r), height)
    else:
        r = width / float(w)
        dim = (width, int(h * r))

    # resize the image
    resized = cv2.resize(image, dim, interpolation=cv2.INTER_AREA)

    # return the resized image
    return resized


def scroll_down(driver):
    total_width = driver.execute_script("return document.body.offsetWidth")
    total_height = driver.execute_script("return document.body.parentNode.scrollHeight")
    viewport_width = driver.execute_script("return document.body.clientWidth")
    viewport_height = driver.execute_script("return window.innerHeight")

    rectangles = []

    i = 0
    while i < total_height:
        ii = 0
        top_height = i + viewport_height

        if top_height > total_height:
            top_height = total_height

        while ii < total_width:
            top_width = ii + viewport_width

            if top_width > total_width:
                top_width = total_width

            rectangles.append((ii, i, top_width, top_height))

            ii = ii + viewport_width

        i = i + viewport_height

    previous = None
    part = 0

    for rectangle in rectangles:
        if not previous is None:
            driver.execute_script("window.scrollTo({0}, {1})".format(rectangle[0], rectangle[1]))
            time.sleep(0.5)
        # time.sleep(0.2)

        if rectangle[1] + viewport_height > total_height:
            offset = (rectangle[0], total_height - viewport_height)
        else:
            offset = (rectangle[0], rectangle[1])

        previous = rectangle

    return (total_height, total_width)


def take_screenshot(driver, screenshot_name="file.png"):
    scroll(driver)
    time.sleep(2)
    height = driver.execute_script("return document.body.parentNode.scrollHeight")
    driver.set_window_size(1920, height)
    time.sleep(2)
    driver.save_screenshot(screenshot_name)
    return driver


def save_page_by_selenium(driver, path):
    f = codecs.open(path, "w", "utf−8")
    h = driver.page_source
    f.write(h)


def save_page_by_singlefile(url, path, is_has_hiddern_element=False):
    # single-file http://web.archive.org/web/20201030104015/https://hiking.biji.co/ web.archive.org.html --browser-executable-path "D:\Program Files\Google\Chrome\Application\chrome.exe"
    os.system('chcp 65001')
    singlefile_bin_path = "D:\portable/node-v18.14.0-win-x64/node_modules/single-file-cli/single-file-kuma.bat"
    command = f"{singlefile_bin_path}  \"{url}\" \"{path}\" --back-end puppeteer --browser-executable-path \"C:/Program Files/Google/Chrome/Application/chrome.exe\" --browser-load-max-time=240000 --browser-debug=false --browser-wait-until-fallback=true --browser-width=1920 --browser-height=1080"
    if is_has_hiddern_element:
        command += " --remove-hidden-elements=false"
    print(command)
    result = ""
    timeout = 300
    try:
        result = subprocess.run(command, timeout=timeout)
    except subprocess.TimeoutExpired:
        logger.error(f'Timeout for {command} ({timeout}s) expired')
    logger.info(f"result: {result}")

def isNaN(string):
    return string != string

def download_url_image(urls, html_folder, image_folder, imgname_2_url_map, names):
    driver = init_driver()
    for i, url in enumerate(urls):
        print(f"process i: {i}, url: {url}")
        if url == "" or isNaN(url):
            continue
        make_directory(f"{html_folder}/{i}")
        selenium_html_name_path = f"{html_folder}/{i}/{names[i]}_selenium.html"
        singlefile_html_name_path = f"{html_folder}/{i}/{names[i]}_singlefile.html"
        singlefile_bs4_html_name_path = f"{html_folder}/{i}/{names[i]}_singlefile_bs4.html"
        html_path = f"{html_folder}/{i}.html"
        singlefile_html_with_hidden_elements_name_path = f"{html_folder}/{i}/{names[i]}_singlefile_with_hidden_elements.html"
        singlefile_with_hidden_elements_bs4_html_name_path = f"{html_folder}/{i}/{names[i]}_singlefile_with_hidden_elements_bs4.html"
        with_hidden_elements_html_path = f"{html_folder}/{i}_with_hidden_elements.html"
        png_image_path = f"{image_folder}/{names[i]}.png"
        jpg_image_path = f"{image_folder}/{names[i]}.jpg"

        imgname_2_url_map[f"{i}.jpg"] = url
        with open(IMGNAME_2_URL_PICKLE_NAME, 'wb') as handle:
            pickle.dump(imgname_2_url_map, handle, protocol=pickle.HIGHEST_PROTOCOL)

        index_jpg_image_path = f"{image_folder}/{i}.jpg"
        if not os.path.exists(index_jpg_image_path) and os.path.exists(jpg_image_path):
            shutil.copyfile(jpg_image_path, index_jpg_image_path)
        # if path.exists(png_image_path):
        #     logger.info(f"exist {i} skip url: {url}")
        #     continue
        driver = open_url(driver=driver, url=url, selenium_html_name_path=selenium_html_name_path,
                          singlefile_html_name_path=singlefile_html_name_path,
                          singlefile_bs4_html_name_path=singlefile_bs4_html_name_path,
                          singlefile_html_with_hidden_elements_name_path=singlefile_html_with_hidden_elements_name_path,
                          singlefile_with_hidden_elements_bs4_html_name_path=singlefile_with_hidden_elements_bs4_html_name_path,
                          html_path=html_path,with_hidden_elements_html_path=with_hidden_elements_html_path
                          , image_path=png_image_path)
        if (not path.exists(jpg_image_path) and path.exists(png_image_path)):
            convert_png_to_jpg(png_image_path, jpg_image_path)

    driver.close()


def write_excel(output_path, imgname_2_url_map, image_folder, is_with_image=False):
    if not imgname_2_url_map:
        # with codecs.open(IMGNAME_2_URL_PICKLE_NAME, "rb") as rf:
        #     imgname_2_url_map = pickle.load(rf)
        with open(IMGNAME_2_URL_PICKLE_NAME, 'rb') as handle:
            imgname_2_url_map = pickle.load(handle)

    workbook = Workbook()
    worksheet = workbook.active

    # images list
    images = []
    for filename in natsorted(glob.glob(f"{image_folder}/*.jpg")):
        images.append(filename)

    # resize cells
    worksheet['A1'] = "Label"
    worksheet['B1'] = "URL"
    worksheet['C1'] = "ImageName"
    if is_with_image:
        worksheet['D1'] = "Image"
    for row in range(2, len(images) + 1):

        # URL column
        for col in range(2, 3):
            col_letter = get_column_letter(col)
            worksheet.column_dimensions[col_letter].width = 50
        # ImageName column
        for col in range(3, 4):
            col_letter = get_column_letter(col)
            worksheet.column_dimensions[col_letter].width = 30

        if is_with_image:
            # Image column
            for col in range(4, 5):
                worksheet.row_dimensions[row].height = 230
                col_letter = get_column_letter(col)
                worksheet.column_dimensions[col_letter].width = 1000

    for index, image in reversed(list(enumerate(images))):

        cell_position = index + 2
        if is_with_image:
            img = OpenpyxlImage(image)
            worksheet.add_image(img, anchor='D' + str(cell_position))
            logger.info(index, image)

        image_name = os.path.basename(image)
        url = "None"
        if (image_name in imgname_2_url_map):
            url = imgname_2_url_map[image_name]
        worksheet.cell(row=cell_position, column=2, value=url)
        worksheet.cell(row=cell_position, column=3, value=image_name)

    # save workbook
    workbook.save(output_path)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Dataset prepare')
    parser.add_argument('--config_file', default='system_from_homepage_dataset_save_page.config',
                        help='Configuration File')
    args = parser.parse_args()
    configs = Configure(config_file=args.config_file)
    fold_check(configs)
    logger = get_logger(log_dir=configs.log_dir)

    # ref_excel_path = "google_search_dataset/pick_rank/google_search_rank_train2.xlsx"
    # output_path = "EventSourcePageURLs.xlsx"
    ref_excel_path = "./dataset_google_search/generate_dataset/pick_rank/google_search_rank_train.xlsx"
    output_path = "/dataset_google_search/generate_dataset/EventSourcePageURLs_withoutimage.xlsx"

    imgname_2_url_map = {}

    # names = ["www.ncu.edu.tw"]
    # urls = ["http://web.archive.org/web/20201204112130/https://www.ncu.edu.tw/"]

    # names = ["www.aheritage.tw"]
    # urls = ["http://web.archive.org/web/20201021040555/http://www.aheritage.tw/"]

    # 87
    # names = ["hiking.biji.co"]
    # urls = ["http://web.archive.org/web/20201030104015/https://hiking.biji.co/"]
    # cache = urls[0]
    # cache_prefix_index = find_nth(cache, "http", 2)
    # cache_prefix = cache[0:cache_prefix_index]
    # print(cache_prefix)
    # cache_relevance_prefix_index = find_nth(cache_prefix, "/web/", 1)
    # cache_relevance_prefix = cache_prefix[cache_relevance_prefix_index:]
    # print(cache_relevance_prefix)

    names, urls = read_urls(ref_excel_path)
    # save_page_by_singlefile("http://web.archive.org/web/20201030104015/https://hiking.biji.co/",
    #                         "D:/tmp/20230203/web.archive.org.html")
    download_url_image(urls=urls, html_folder=configs.html_dir, image_folder=configs.img_dir,
                       imgname_2_url_map=imgname_2_url_map, names=names)
    # write_excel(output_path=output_path, imgname_2_url_map=imgname_2_url_map, image_folder=configs.img_dir)
